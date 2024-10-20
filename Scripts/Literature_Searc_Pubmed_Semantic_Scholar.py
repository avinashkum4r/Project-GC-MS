import requests
import time
import mysql.connector
import os
from openpyxl import Workbook, load_workbook

# Excel file to store results
file_name = "output\\relevant_research_papers.xlsx"

# Initialize Excel workbook and sheet if the file does not exist
try:
    wb = load_workbook(file_name)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["Source", "Keyword", "Year/Date", "Title", "Authors", "URL", "Paper ID"])
    wb.save(file_name)

# Keywords for literature search
keywords = [
    "lead AND corn AND gcms AND pb",
    "lead AND plants AND gcms",
    "lead AND gcms AND toxicity",
    "lead AND gcms AND metabolites AND stress",
    "lead AND corn AND gcms"
]

# Get ENV values for MySQL connection
sql_host = os.getenv('MYSQL_HOST')
sql_user = os.getenv('MYSQL_USER')
sql_password = os.getenv('MYSQL_PASSWORD')
sql_db = os.getenv('MYSQL_GCMS_DB')

# MySQL connection details
db_config = {
    'host': sql_host,
    'user': sql_user,
    'password': sql_password,
    'database': sql_db
}

# Retry and wait time settings
MAX_RETRIES = 3
WAIT_TIME = 5  # seconds


# Helper function to write results to Excel
def write_to_excel(source, keyword, title, year, authors, url, paper_id):
    ws.append([source, keyword, year, title, authors, url, paper_id])
    wb.save(file_name)


# Function to insert data into MySQL database
def insert_into_mysql(data):
    try:
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor()
            query = """INSERT INTO gcms.research_papers (Source, Keyword, Title, Year, Authors, URL, PaperID)
                       VALUES (%s, %s, %s, %s, %s, %s, %s)"""
            cursor.executemany(query, data)
            connection.commit()
            print(f"{cursor.rowcount} records inserted successfully into MySQL.")
    except mysql.connector.Error as e:
        print(f"Error connecting to MySQL: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
            print("MySQL connection closed.")


# PubMed Search
def fetch_pubmed_details(pmid_list, keyword):
    pmid_str = ",".join(pmid_list)
    url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi?db=pubmed&id={pmid_str}&retmode=json"
    retries = 0
    while retries < MAX_RETRIES:
        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()
            docs = data.get("result", {})
            mysql_data = []
            for pmid in pmid_list:
                doc = docs.get(pmid, {})
                title = doc.get("title", "N/A")
                year = doc.get("pubdate", "N/A")

                # Check if year is valid and trim it to 4 digits if necessary
                if isinstance(year, str) and len(year) > 4:
                    year = year[:4]  # Take only the first 4 characters
                elif not year or not year.isdigit():
                    year = "N/A"  # Default value if year is not valid

                authors = ", ".join([author.get("name", "N/A") for author in doc.get("authors", [])])
                article_url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
                write_to_excel("PubMed", keyword, title, year, authors, article_url, pmid)
                mysql_data.append(("PubMed", keyword, title, year, authors, article_url, pmid))
            insert_into_mysql(mysql_data)
            break
        except Exception as e:
            print(f"Error fetching PubMed details: {e}")
            retries += 1
            time.sleep(WAIT_TIME)


def search_pubmed(keyword):
    url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?retmode=json&db=pubmed&term={keyword}&retmax=5"
    retries = 0
    while retries < MAX_RETRIES:
        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()
            pmid_list = data.get("esearchresult", {}).get("idlist", [])
            if pmid_list:
                fetch_pubmed_details(pmid_list, keyword)
            break
        except Exception as e:
            print(f"Error fetching PubMed IDs: {e}")
            retries += 1
            time.sleep(WAIT_TIME)


# Semantic Scholar Search
def search_semantic_scholar(keyword):
    url = f"https://api.semanticscholar.org/graph/v1/paper/search?query={keyword}&fields=title,authors,year,url&limit=5"
    retries = 0
    while retries < MAX_RETRIES:
        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()
            papers = data.get("data", [])
            mysql_data = []
            for paper in papers:
                title = paper.get("title", "N/A")
                year = paper.get("year", "N/A")
                authors = ", ".join([author.get("name", "N/A") for author in paper.get("authors", [])])
                paper_url = paper.get("url", "N/A")
                paper_id = paper.get("paperId", "N/A")
                write_to_excel("Semantic Scholar", keyword, title, year, authors, paper_url, paper_id)
                mysql_data.append(("Semantic Scholar", keyword, title, year, authors, paper_url, paper_id))
            insert_into_mysql(mysql_data)
            break
        except Exception as e:
            print(f"Error fetching Semantic Scholar papers: {e}")
            retries += 1
            time.sleep(WAIT_TIME)


# Main search loop through keywords
for keyword in keywords:
    print(f"Searching PubMed for keyword: {keyword}")
    search_pubmed(keyword)

    print(f"Searching Semantic Scholar for keyword: {keyword}")
    search_semantic_scholar(keyword)

print("Search completed across all sources.")
